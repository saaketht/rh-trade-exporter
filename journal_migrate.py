#!/usr/bin/env python3
"""One-time journal migration after the order_id regroup.

The Trade Log in spy_0dte_journal.xlsx was built from the legacy fragmented CSV.
After regroup_trades.py collapsed broker fragmentation in spy_trades.csv, the
journal's Qtys / row counts / Group IDs no longer match, so a plain journal_sync
would double-append. This rebuilds the Trade Log from the clean CSV and carries
the manual columns (Setup / Trigger / Exit Reason / Rules / Notes — AA-AE)
forward by position.

Manual notes describe a *position*, so they're keyed by (Date, Strike, Type) and
re-attached to that position's new row(s). When a position collapsed from several
fragments into one row, the note follows it. When a position legitimately stays
multiple rows (a scale-out), the note attaches to the earliest-exit row (flagged).

Dashboard and Setup Analysis sheets are never touched. Dry-run by default;
--apply backs up to *.prefragfix.bak then writes in place. Aborts if any manual
note cannot be placed (no silent loss).

  python journal_migrate.py            # dry run — prints the full note mapping
  python journal_migrate.py --apply    # back up + rewrite the Trade Log
"""
import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

import journal_sync as js

MANUAL = {"setup": 27, "trigger": 28, "reason": 29, "rules": 30, "notes": 31}
SHEET = js.SHEET_NAME
ENTRY_TOL = 90  # seconds — fragments share an entry instant; re-entries are minutes apart


def _pos(date, strike, opt_type):
    return (date.isoformat() if date else None,
            float(strike) if strike is not None else None,
            str(opt_type).strip().lower()[:1] if opt_type else None)


def _secs(t):
    return None if t is None else t.hour * 3600 + t.minute * 60 + t.second


def read_notes(ws):
    """Collect each manually-noted Trade Log row individually (NOT merged by
    position — distinct re-entries on the same strike/day carry different notes).
    Returns a list of {pos, entry, exit, vals}."""
    out = []
    for r in range(2, ws.max_row + 1):
        date = ws.cell(r, js.COL["date"]).value
        if date is None:
            continue
        vals = {name: ws.cell(r, ci).value for name, ci in MANUAL.items()
                if ws.cell(r, ci).value not in (None, "", " ")}
        if not vals:
            continue
        d = date.date() if hasattr(date, "date") else date
        out.append({
            "pos": _pos(d, ws.cell(r, js.COL["strike"]).value, ws.cell(r, js.COL["type"]).value),
            "entry": _secs(ws.cell(r, js.COL["entry_time"]).value),
            "exit": _secs(ws.cell(r, js.COL["exit_time"]).value),
            "vals": vals,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--journal", default=str(js.DEFAULT_JOURNAL))
    ap.add_argument("--csv", default=str(js.DEFAULT_CSV))
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    jpath, cpath = Path(args.journal), Path(args.csv)
    wb = load_workbook(jpath, keep_vba=False, data_only=False)
    ws = wb[SHEET]
    old_last = ws.max_row

    noted = read_notes(ws)
    print(f"📖 {jpath.name}: {old_last - 1} data rows, {len(noted)} noted rows")

    trades, _ = js.load_csv_trades(cpath, set())
    print(f"📄 {cpath.name}: {len(trades)} clean trades")

    # candidate new rows per position: (row, entry_secs, exit_secs)
    pos_rows = {}
    for i, t in enumerate(trades):
        pos_rows.setdefault(_pos(t["date"], t["strike"], t["type"]), []).append(
            (2 + i, _secs(t["entry_time"]), _secs(t["exit_time"])))

    # match each noted row to the new row with the same (Date,Strike,Type) and
    # nearest entry time (within tolerance), tie-broken by exit time. Fragments
    # (same entry instant) land on the merged row; re-entries stay distinct.
    row_vals = {}        # new_row -> {col: value}
    conflicts, unplaced = [], []
    for rec in noted:
        cands = pos_rows.get(rec["pos"], [])
        best, best_key = None, None
        for (row, en, ex) in cands:
            if rec["entry"] is not None and en is not None and abs(en - rec["entry"]) > ENTRY_TOL:
                continue
            ed = abs((en or 0) - (rec["entry"] or 0))
            xd = abs((ex or 0) - (rec["exit"] or 0))
            if best_key is None or (ed, xd) < best_key:
                best_key, best = (ed, xd), row
        if best is None:
            unplaced.append(rec)
            continue
        bucket = row_vals.setdefault(best, {})
        for name, v in rec["vals"].items():
            if name in bucket and str(bucket[name]).strip() != str(v).strip():
                conflicts.append((rec["pos"], name, bucket[name], v))
            else:
                bucket.setdefault(name, v)

    print(f"\n── note mapping ──")
    print(f"  noted rows placed: {len(noted) - len(unplaced)}/{len(noted)}  "
          f"-> {len(row_vals)} distinct target rows")
    print(f"  same-row merge conflicts (first kept): {len(conflicts)}")
    for pos, name, a, b in conflicts[:8]:
        print(f"     {pos} {name}: kept {str(a)[:40]!r} over {str(b)[:40]!r}")
    if unplaced:
        print(f"  ⚠ UNPLACED ({len(unplaced)}) — no new row within {ENTRY_TOL}s:")
        for rec in unplaced[:10]:
            print(f"     {rec['pos']} entry={rec['entry']}s vals={list(rec['vals'])}")
        print("  ABORTING — would lose manual work. Investigate before --apply.")
        sys.exit(2)

    if not args.apply:
        print("\n(dry run — pass --apply to back up + rewrite the Trade Log)")
        for row in sorted(row_vals)[:8]:
            t = trades[row - 2]
            print(f"    row {row}  {t['date']} {t['strike']:.0f}{t['type'][:1]} "
                  f"{t['entry_time']} -> {list(row_vals[row])}")
        return

    # ── apply ──
    bak = jpath.with_suffix(".prefragfix.bak.xlsx")
    if not bak.exists():
        shutil.copy(jpath, bak)
        print(f"\n🛟 backed up -> {bak.name}")

    # clear all data cells (cols 1..37) for every old data row
    for r in range(2, old_last + 1):
        for c in range(1, 38):
            ws.cell(r, c).value = None

    # rebuild Trade Log from clean CSV (formulas/formats; AA-AE left blank)
    js.append_rows(ws, trades, start_row=2, last_trade_num=0)

    # re-attach manual notes
    placed = 0
    for target, vals in row_vals.items():
        for name, v in vals.items():
            ws.cell(target, MANUAL[name]).value = v
        placed += 1

    new_last = 1 + len(trades)
    js.extend_table_range(ws, new_last)
    js.strip_stale_dvs(ws)

    wb.save(jpath)
    print(f"✅ wrote {jpath.name}: {len(trades)} rows, {placed} note-positions re-attached")
    print("   (open in Excel to let formulas recompute; Dashboard/Setup Analysis untouched)")


if __name__ == "__main__":
    main()

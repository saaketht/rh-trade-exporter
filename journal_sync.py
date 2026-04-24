#!/usr/bin/env python3
"""
journal_sync.py — Merge hood.py CSV output into spy_0dte_journal.xlsx.

Appends only new trades (dedup by Date + Entry Time + Strike + Type + Qty).
Preserves existing rows, formulas, formatting, and the Setup/Notes columns
(AA-AE) for manual fill by the user.

Default output: spy_0dte_journal_updated.xlsx (original untouched).
Use --in-place to overwrite the original.
Use --fetch to auto-run hood.py for dates after the journal's last row.
"""
from __future__ import annotations

import argparse
import csv
import subprocess
import sys
from copy import copy
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Row height for appended rows — matches the journal's design for multi-line trader notes.
ROW_HEIGHT = 31

# Columns that should be left-aligned (trader notes — variable-length text).
# All other columns get center-aligned. Vertical alignment is center on everything.
#   AA (27): Setup   AB (28): Trigger   AC (29): Exit Reason   AE (31): Notes
# AD (Rules Followed) stays center because its values are Yes/No/Partial.
LEFT_ALIGN_COLS = {27, 28, 29, 31}

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_JOURNAL = SCRIPT_DIR / "spy_0dte_journal.xlsx"
DEFAULT_CSV = SCRIPT_DIR / "outputs" / "spy_trades.csv"
DEFAULT_OUTPUT = SCRIPT_DIR / "spy_0dte_journal_updated.xlsx"
SHEET_NAME = "Trade Log"

# Column indices (1-based, matching openpyxl)
COL = {
    "trade_num": 1, "date": 2, "day": 3, "underlying": 4, "expiry": 5,
    "strike": 6, "type": 7, "qty": 8,
    "open": 9, "high": 10, "low": 11, "close": 12,
    "vwap": 13, "ema8": 14, "trend_aligned": 15,
    "entry_time": 16, "exit_time": 17, "hold_time": 18, "entry_hour": 19,
    "entry_cost": 20, "exit_credit": 21, "pl_dollar": 22, "cum_pl": 23,
    "pl_pct": 24, "win_loss": 25, "is_win": 26,
    "setup": 27, "trigger": 28, "reason": 29, "rules": 30, "notes": 31,
    "vix": 32, "risk": 33, "r_multiple": 34, "delta": 35, "group_id": 36,
    "dte": 37,
}

# Explicit per-column number formats. Row 2 of the journal has junk formats on
# several columns (e.g. col D 'SPY' with mm/dd/yyyy, col G 'Call' with $#,##0.00,
# col N 'N/A' with hh:mm:ss) — copying from row 2 propagates those. This map is
# the source of truth for formats on newly appended rows.
DATE_FMT = "mm/dd/yyyy"
TIME_FMT = "hh:mm:ss"
CURRENCY_FMT = '\\$#,##0.00_);[Red]"($"#,##0.00\\)'
PLAIN_CURRENCY = "\\$#,##0.00"
INT_CURRENCY = "\\$#,##0"
PERCENT_FMT = "0.00%"
TWO_DEC = "0.00"

COLUMN_FORMATS = {
    "trade_num": "General",
    "date": DATE_FMT,
    "day": "General",
    "underlying": "General",
    "expiry": DATE_FMT,
    "strike": "General",
    "type": "General",
    "qty": "0",
    "open": CURRENCY_FMT, "high": CURRENCY_FMT, "low": CURRENCY_FMT, "close": CURRENCY_FMT,
    "vwap": "General",        # now categorical (Above/Below/At/N/A)
    "ema8": "General",        # now categorical
    "trend_aligned": "General",
    "entry_time": TIME_FMT,
    "exit_time": TIME_FMT,
    "hold_time": "h:mm;@",
    "entry_hour": INT_CURRENCY,  # matches existing journal col S formatting
    "entry_cost": PLAIN_CURRENCY,
    "exit_credit": PLAIN_CURRENCY,
    "pl_dollar": PLAIN_CURRENCY,
    "cum_pl": PLAIN_CURRENCY,
    "pl_pct": PERCENT_FMT,
    "win_loss": "General",
    "is_win": "General",
    "setup": "General", "trigger": "General", "reason": "General",
    "rules": "General", "notes": "General",
    "vix": TWO_DEC,
    "risk": "General",
    "r_multiple": TWO_DEC,
    "delta": TWO_DEC,
    "group_id": "General",
    "dte": "General",
}

# Formula templates — row-substituted at write time. These match the patterns
# in row 2 of the existing journal (inspected before implementation).
FORMULAS = {
    "trend_aligned": (
        '=IF(AND(M{r}="N/A", N{r}="N/A"), "N/A", '
        'IF(AND(OR(M{r}="Above", M{r}="At"), OR(N{r}="Above", N{r}="At"), G{r}="Call"), "Yes", '
        'IF(AND(OR(M{r}="Below", M{r}="At"), OR(N{r}="Below", N{r}="At"), G{r}="Put"), "Yes", "No")))'
    ),
    "hold_time": "=MOD(Q{r}-P{r}, 1)",
    "entry_hour": "=HOUR('Trade Log'!$P{r})",
    "pl_dollar": "=U{r}+T{r}",
    "cum_pl": "=SUM($V$2:V{r})",
    "pl_pct": "=IF(T{r}=0,0,V{r}/ABS(T{r}))",
    "win_loss": '=IF(V{r}>0,"WIN",IF(V{r}<0,"LOSS","BE"))',
    "is_win": "=IF('Trade Log'!$Y{r}=\"WIN\", 1, 0)",
    "risk": "=ABS('Trade Log'!$T{r})",
    "r_multiple": '=IF(AG{r}<>"",V{r}/ABS(AG{r}),"")',
}


# ───────────────────────── helpers ─────────────────────────

def parse_csv_date(s: str) -> date | None:
    """CSV dates are M/D/YYYY."""
    if not s or s.strip() == "":
        return None
    try:
        return datetime.strptime(s.strip(), "%m/%d/%Y").date()
    except ValueError:
        try:
            return datetime.strptime(s.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None


def parse_csv_time(s: str) -> time | None:
    """CSV times are HH:MM:SS."""
    if not s or s.strip() == "":
        return None
    try:
        return datetime.strptime(s.strip(), "%H:%M:%S").time()
    except ValueError:
        return None


def parse_number(s):
    """Parse a CSV cell as float, or None if blank/NaN."""
    if s is None:
        return None
    s = str(s).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(s):
    v = parse_number(s)
    return int(v) if v is not None else None


_TREND_VALUES = {"above", "below", "at", "n/a"}


def _coerce_trend(s):
    """Normalize VWAP/8 EMA CSV cell to one of 'Above'/'Below'/'At'/'N/A'.
    Any unknown value (including legacy numeric) maps to 'N/A'."""
    if s is None:
        return "N/A"
    v = str(s).strip()
    if v == "" or v.lower() == "nan":
        return "N/A"
    if v.lower() in _TREND_VALUES:
        return v.capitalize() if v.lower() != "n/a" else "N/A"
    return "N/A"


def dedup_key(trade_date, entry_time_obj, strike, opt_type, qty):
    """Composite key: (date iso, time iso, strike float, type lowercase, qty int).
    Robust against minor format variations."""
    return (
        trade_date.isoformat() if trade_date else None,
        entry_time_obj.isoformat() if entry_time_obj else None,
        float(strike) if strike is not None else None,
        str(opt_type).strip().lower() if opt_type else None,
        int(qty) if qty is not None else None,
    )


# ───────────────────────── journal I/O ─────────────────────────

def load_journal(path: Path):
    """Return (workbook, sheet, last_row, existing_keys_set)."""
    if not path.exists():
        print(f"❌ Journal not found: {path}", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, keep_vba=False)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not in {path}", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET_NAME]
    last_row = ws.max_row

    existing = set()
    max_date = None
    for r in range(2, last_row + 1):
        d = ws.cell(r, COL["date"]).value
        t = ws.cell(r, COL["entry_time"]).value
        strike = ws.cell(r, COL["strike"]).value
        opt_type = ws.cell(r, COL["type"]).value
        qty = ws.cell(r, COL["qty"]).value

        if isinstance(d, datetime):
            d = d.date()
        if isinstance(t, datetime):
            t = t.time()

        if d and isinstance(d, date) and (max_date is None or d > max_date):
            max_date = d

        if d and t and strike is not None and opt_type and qty is not None:
            existing.add(dedup_key(d, t, strike, opt_type, qty))

    return wb, ws, last_row, existing, max_date


def load_csv_trades(csv_path: Path, existing_keys: set) -> list[dict]:
    if not csv_path.exists():
        print(f"❌ CSV not found: {csv_path}", file=sys.stderr)
        sys.exit(1)

    new_trades = []
    skipped = 0
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            trade_date = parse_csv_date(row.get("Date", ""))
            entry_t = parse_csv_time(row.get("Entry Time", ""))
            strike = parse_number(row.get("Strike"))
            opt_type = row.get("Type", "")
            qty = parse_int(row.get("Qty"))

            if not (trade_date and entry_t and strike is not None and opt_type and qty is not None):
                continue

            key = dedup_key(trade_date, entry_t, strike, opt_type, qty)
            if key in existing_keys:
                skipped += 1
                continue

            new_trades.append({
                "date": trade_date,
                "day": row.get("Day", "").strip(),
                "symbol": row.get("Symbol", "").strip(),
                "expiry": parse_csv_date(row.get("Expiry Date", "")),
                "strike": strike,
                "type": opt_type.strip(),
                "qty": qty,
                "open": parse_number(row.get("Asset Open")),
                "high": parse_number(row.get("Asset High")),
                "low": parse_number(row.get("Asset Low")),
                "close": parse_number(row.get("Asset Close")),
                # VWAP / 8 EMA are categorical: "Above" / "Below" / "At" / "N/A".
                # Older CSVs may contain numeric values — coerce any non-category to "N/A".
                "vwap": _coerce_trend(row.get("VWAP")),
                "ema8": _coerce_trend(row.get("8 EMA")),
                "entry_time": entry_t,
                "exit_time": parse_csv_time(row.get("Exit Time", "")),
                "entry_cost": parse_number(row.get("Entry Cost")),
                "exit_credit": parse_number(row.get("Exit Credit")),
                "vix": parse_number(row.get("VIX")),
                "delta": parse_number(row.get("Delta")),
                "group_id": (row.get("Group ID", "") or "").strip(),
                "dte": parse_int(row.get("DTE")),
            })

    return new_trades, skipped


# ───────────────────────── writing ─────────────────────────

_COL_IDX_TO_KEY = {v: k for k, v in COL.items()}


# Columns whose fill + border should be carried from row 2 (AE = Notes/Lessons
# has pale yellow highlighting the user wants preserved on new rows).
STYLE_CARRY_COLS = {31}  # AE


def apply_column_format(target_cell, col_idx: int, source_row: int = 2, ws=None):
    """Apply the explicit per-column number format + alignment. Carries fill/border
    from `source_row` for columns in STYLE_CARRY_COLS so decorative styling (e.g.
    the yellow Notes column) persists on appended rows."""
    key = _COL_IDX_TO_KEY.get(col_idx)
    if key and key in COLUMN_FORMATS:
        target_cell.number_format = COLUMN_FORMATS[key]
    horizontal = "left" if col_idx in LEFT_ALIGN_COLS else "center"
    target_cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

    if ws is not None and col_idx in STYLE_CARRY_COLS:
        src = ws.cell(row=source_row, column=col_idx)
        target_cell.fill = copy(src.fill)
        target_cell.border = copy(src.border)


def append_rows(ws, trades: list[dict], start_row: int, last_trade_num: int):
    """Write trades to the sheet starting at start_row.

    GUARANTEE: every write is at row >= start_row. Never touches rows < start_row.
    """
    for i, t in enumerate(trades):
        r = start_row + i
        trade_num = last_trade_num + 1 + i

        # Build (col_idx, value) list — hardcoded data first, then formulas.
        # Any NaN/None becomes a blank cell.
        writes = [
            (COL["trade_num"], trade_num),
            (COL["date"], t["date"]),
            (COL["day"], t["day"] or None),
            (COL["underlying"], t["symbol"] or None),
            (COL["expiry"], t["expiry"]),
            (COL["strike"], t["strike"]),
            (COL["type"], t["type"] or None),
            (COL["qty"], t["qty"]),
            (COL["open"], t["open"]),
            (COL["high"], t["high"]),
            (COL["low"], t["low"]),
            (COL["close"], t["close"]),
            (COL["vwap"], t["vwap"]),
            (COL["ema8"], t["ema8"]),
            (COL["trend_aligned"], FORMULAS["trend_aligned"].format(r=r)),
            (COL["entry_time"], t["entry_time"]),
            (COL["exit_time"], t["exit_time"]),
            (COL["hold_time"], FORMULAS["hold_time"].format(r=r)),
            (COL["entry_hour"], FORMULAS["entry_hour"].format(r=r)),
            (COL["entry_cost"], t["entry_cost"]),
            (COL["exit_credit"], t["exit_credit"]),
            (COL["pl_dollar"], FORMULAS["pl_dollar"].format(r=r)),
            (COL["cum_pl"], FORMULAS["cum_pl"].format(r=r)),
            (COL["pl_pct"], FORMULAS["pl_pct"].format(r=r)),
            (COL["win_loss"], FORMULAS["win_loss"].format(r=r)),
            (COL["is_win"], FORMULAS["is_win"].format(r=r)),
            # AA-AE (setup, trigger, reason, rules, notes) — intentionally skipped.
            (COL["vix"], t["vix"]),
            (COL["risk"], FORMULAS["risk"].format(r=r)),
            (COL["r_multiple"], FORMULAS["r_multiple"].format(r=r)),
            (COL["delta"], t["delta"]),
            (COL["group_id"], t["group_id"] or None),
            (COL["dte"], t["dte"]),
        ]

        written_cols = {c for c, _ in writes}
        for col_idx, value in writes:
            cell = ws.cell(row=r, column=col_idx, value=value)
            apply_column_format(cell, col_idx, ws=ws)

        # Pre-set alignment on manual-fill columns (AA-AE) so the user gets the
        # right formatting when they type. No value written — cells stay empty.
        for col_idx in range(1, 38):
            if col_idx not in written_cols:
                apply_column_format(ws.cell(row=r, column=col_idx), col_idx, ws=ws)

        ws.row_dimensions[r].height = ROW_HEIGHT


# Columns where stale data validations leaked onto rows 150+ due to prior row
# shifts in the source journal. We strip DVs from these columns on rows ≥ 150
# so appended rows stay clean. AA (27, Setup) is intentionally NOT stripped —
# user may add a dropdown there later.
DV_STRIP_COLS = {3, 5, 11, 12, 28}  # C, E, K, L, AB
DV_STRIP_FROM_ROW = 150


def strip_stale_dvs(ws):
    """Remove data validations from DV_STRIP_COLS for rows ≥ DV_STRIP_FROM_ROW.
    Rewrites each DV's sqref to exclude those ranges; drops DVs left empty.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.cell_range import CellRange

    dv_list = ws.data_validations.dataValidation
    keep = []
    stripped_total = 0
    for dv in dv_list:
        new_ranges = []
        for cr in dv.sqref.ranges:
            col_min, col_max = cr.min_col, cr.max_col
            row_min, row_max = cr.min_row, cr.max_row
            overlap_cols = [c for c in range(col_min, col_max + 1) if c in DV_STRIP_COLS]
            if not overlap_cols or row_max < DV_STRIP_FROM_ROW:
                new_ranges.append(cr)
                continue
            # Rebuild ranges per column, clipping out rows ≥ DV_STRIP_FROM_ROW for stripped cols.
            for c in range(col_min, col_max + 1):
                col_letter = get_column_letter(c)
                if c in DV_STRIP_COLS:
                    if row_min < DV_STRIP_FROM_ROW:
                        new_ranges.append(CellRange(
                            f"{col_letter}{row_min}:{col_letter}{DV_STRIP_FROM_ROW - 1}"
                        ))
                    stripped_total += 1
                else:
                    new_ranges.append(CellRange(f"{col_letter}{row_min}:{col_letter}{row_max}"))
        if new_ranges:
            dv.sqref = " ".join(str(r) for r in new_ranges)
            keep.append(dv)
    ws.data_validations.dataValidation = keep
    if stripped_total:
        print(f"   Stripped stale DVs from {stripped_total} column-range(s) on rows ≥ {DV_STRIP_FROM_ROW}")


def extend_table_range(ws, new_last_row: int):
    """Expand the Trade Log Excel table to include the new last row.
    Excel tables need their `ref` updated when rows are appended outside the original range.
    """
    for name in list(ws.tables.keys()):
        tbl = ws.tables[name]
        # Parse existing ref like 'A1:AK149' → keep start, update end row.
        ref = tbl.ref
        if ":" not in ref:
            continue
        start, end = ref.split(":")
        # Strip digits from end to isolate column letters.
        end_col = "".join(ch for ch in end if ch.isalpha())
        tbl.ref = f"{start}:{end_col}{new_last_row}"


# ───────────────────────── fetch integration ─────────────────────────

def run_hood_fetch(after_date: date | None) -> None:
    """Subprocess hood.py with --after-date <max_journal_date>. Abort on failure."""
    cmd = [sys.executable, str(SCRIPT_DIR / "hood.py"), "--symbol", "SPY"]
    if after_date:
        cmd += ["--after-date", after_date.isoformat()]
    print(f"🪝 Running: {' '.join(cmd)}")
    result = subprocess.run(cmd, cwd=SCRIPT_DIR)
    if result.returncode != 0:
        print(f"❌ hood.py failed (exit {result.returncode}). Journal NOT modified.", file=sys.stderr)
        sys.exit(result.returncode)


# ───────────────────────── safety diff ─────────────────────────

def assert_existing_rows_untouched(orig_path: Path, new_path: Path, last_row: int):
    """Read rows 1..last_row from both files. Abort if any cell differs."""
    wb_a = load_workbook(orig_path, keep_vba=False, data_only=False)
    wb_b = load_workbook(new_path, keep_vba=False, data_only=False)
    ws_a = wb_a[SHEET_NAME]
    ws_b = wb_b[SHEET_NAME]

    max_col = max(ws_a.max_column, ws_b.max_column)
    diffs = []
    for r in range(1, last_row + 1):
        for c in range(1, max_col + 1):
            va = ws_a.cell(r, c).value
            vb = ws_b.cell(r, c).value
            if va != vb:
                diffs.append((r, get_column_letter(c), va, vb))

    if diffs:
        print(f"❌ SAFETY CHECK FAILED: {len(diffs)} cells in existing rows changed!", file=sys.stderr)
        for r, col, va, vb in diffs[:10]:
            print(f"   {col}{r}: {va!r} → {vb!r}", file=sys.stderr)
        sys.exit(2)
    print(f"✓ Safety check: all {last_row} existing rows unchanged across {max_col} columns.")


# ───────────────────────── main ─────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--journal", type=Path, default=DEFAULT_JOURNAL)
    ap.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument("--in-place", action="store_true",
                    help="Overwrite the original journal (default: write to _updated.xlsx).")
    ap.add_argument("--fetch", action="store_true",
                    help="Auto-run hood.py for dates after the journal's last row before appending.")
    args = ap.parse_args()

    out_path = args.journal if args.in_place else args.output

    # Peek at journal to get last_date for --fetch, before full load.
    if args.fetch:
        _wb, _ws, _lr, _keys, peek_max_date = load_journal(args.journal)
        _wb.close()
        run_hood_fetch(peek_max_date)

    # Real load.
    wb, ws, last_row, existing_keys, max_date = load_journal(args.journal)
    print(f"📖 Journal: {args.journal.name}  ({last_row} rows, last trade {max_date})")
    print(f"   Existing dedup keys: {len(existing_keys)}")

    # Last Trade # in col A.
    last_trade_num = 0
    for r in range(2, last_row + 1):
        v = ws.cell(r, COL["trade_num"]).value
        if isinstance(v, (int, float)):
            last_trade_num = max(last_trade_num, int(v))

    new_trades, skipped = load_csv_trades(args.csv, existing_keys)
    print(f"📄 CSV: {args.csv.name}  →  {len(new_trades)} new, {skipped} duplicates (skipped)")

    if not new_trades:
        print("✓ Nothing to append. Journal is up to date.")
        return

    append_rows(ws, new_trades, start_row=last_row + 1, last_trade_num=last_trade_num)
    extend_table_range(ws, new_last_row=last_row + len(new_trades))
    strip_stale_dvs(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"💾 Wrote: {out_path}  (+{len(new_trades)} rows, last Trade #{last_trade_num + len(new_trades)})")

    # Safety diff — only meaningful when original and output are different files.
    if not args.in_place:
        assert_existing_rows_untouched(args.journal, out_path, last_row)
    else:
        print("   (safety diff skipped in --in-place mode)")


if __name__ == "__main__":
    main()

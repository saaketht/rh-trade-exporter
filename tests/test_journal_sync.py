"""Tests for journal_sync.py — pure logic + a synthetic-workbook smoke test.

The full Excel append flow is integration-heavy (real journal, real CSV), so we
synthesize a minimal workbook in-memory and assert the append loop's invariants:
  - existing rows untouched
  - new rows populated with expected values + formulas
  - stale DVs stripped on target columns, preserved on AA
  - row height + alignment applied
"""
from __future__ import annotations

from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

import journal_sync as js


# ───────────────────────── dedup_key ─────────────────────────

class TestDedupKey:
    def test_normalizes_type_case_and_strike_type(self):
        k1 = js.dedup_key(date(2026, 4, 23), time(9, 30), 580, "CALL", 2)
        k2 = js.dedup_key(date(2026, 4, 23), time(9, 30), 580.0, "call", "2")
        assert k1 == k2

    def test_none_fields_pass_through(self):
        k = js.dedup_key(None, None, None, None, None)
        assert k == (None, None, None, None, None)


# ───────────────────────── _coerce_trend ─────────────────────────

class TestCoerceTrend:
    @pytest.mark.parametrize("raw,expected", [
        ("Above", "Above"),
        ("below", "Below"),
        ("AT", "At"),
        ("N/A", "N/A"),
        ("n/a", "N/A"),
        ("", "N/A"),
        (None, "N/A"),
        ("nan", "N/A"),
        ("580.23", "N/A"),        # legacy numeric → N/A
        ("something-weird", "N/A"),
    ])
    def test_coerce(self, raw, expected):
        assert js._coerce_trend(raw) == expected


# ───────────────────────── parse helpers ─────────────────────────

class TestParseHelpers:
    def test_parse_csv_date_mdy(self):
        assert js.parse_csv_date("4/23/2026") == date(2026, 4, 23)

    def test_parse_csv_date_iso_fallback(self):
        assert js.parse_csv_date("2026-04-23") == date(2026, 4, 23)

    def test_parse_csv_date_blank(self):
        assert js.parse_csv_date("") is None

    def test_parse_csv_time(self):
        assert js.parse_csv_time("09:30:15") == time(9, 30, 15)

    def test_parse_number_nan(self):
        assert js.parse_number("nan") is None
        assert js.parse_number("") is None
        assert js.parse_number("3.14") == 3.14


# ───────────────────────── DV strip ─────────────────────────

def _seed_workbook_with_dvs():
    """Minimal workbook with DVs across ranges that mimic the real journal's
    stale DVs (shifted onto rows ≥150 for C, E, K, L, AB) plus a legitimate
    DV on AA (Setup) that must be preserved."""
    wb = Workbook()
    ws = wb.active
    ws.title = js.SHEET_NAME

    # Simulated stale DVs (spans both pre-150 and post-150 rows).
    stale = DataValidation(type="list", formula1='"Call,Put"')
    stale.add("E2:E1043")     # E crosses 150 → post-150 stripped
    stale.add("K150:K1017")   # K entirely post-150 → fully stripped
    stale.add("AB150:AB1043") # AB entirely post-150 → fully stripped
    ws.add_data_validation(stale)

    # Legitimate AA (27) DV — must survive untouched.
    setup_dv = DataValidation(type="list", formula1='"Breakout,Reversal"')
    setup_dv.add("AA2:AA500")
    ws.add_data_validation(setup_dv)
    return wb, ws


class TestStripStaleDvs:
    def test_strips_post_150_ranges(self):
        wb, ws = _seed_workbook_with_dvs()
        js.strip_stale_dvs(ws)

        # Collect surviving ranges as strings for assertions.
        surviving = []
        for dv in ws.data_validations.dataValidation:
            for cr in dv.sqref.ranges:
                surviving.append(str(cr))

        # E2:E149 should survive (pre-150 slice of the old E2:E1043 range).
        assert any("E2:E149" in s for s in surviving), surviving
        # Nothing on rows ≥150 for C, E, K, L, AB.
        for s in surviving:
            if any(s.startswith(col) for col in ("C", "E", "K", "L", "AB")):
                # Extract the row range numbers; must not include any ≥150.
                _, end = s.split(":")
                end_row = int("".join(ch for ch in end if ch.isdigit()))
                assert end_row < 150, f"leaked stale DV: {s}"

    def test_preserves_aa_dv(self):
        wb, ws = _seed_workbook_with_dvs()
        js.strip_stale_dvs(ws)

        surviving = [str(cr)
                     for dv in ws.data_validations.dataValidation
                     for cr in dv.sqref.ranges]
        assert any(s.startswith("AA") for s in surviving), \
            "AA (Setup) DV should have been preserved"


# ───────────────────────── append_rows smoke test ─────────────────────────

def _seed_journal_with_row2():
    """Make a workbook with a header row + one data row (row 2), so row-2
    format/fill-carry behavior can be exercised without requiring the real file."""
    wb = Workbook()
    ws = wb.active
    ws.title = js.SHEET_NAME
    # Minimal header (values don't matter for these tests).
    for c in range(1, 38):
        ws.cell(row=1, column=c, value=f"H{c}")
    # Seed row 2 with placeholder cells (openpyxl needs cells to exist for copy).
    for c in range(1, 38):
        ws.cell(row=2, column=c, value=None)
    return wb, ws


def _sample_trade(**overrides):
    base = dict(
        date=date(2026, 4, 23), day="Thu", symbol="SPY",
        expiry=date(2026, 4, 23), strike=580.0, type="Call", qty=2,
        open=580.5, high=582.1, low=579.8, close=581.2,
        vwap="Above", ema8="Below",
        entry_time=time(9, 30, 0), exit_time=time(9, 45, 0),
        entry_cost=-500.0, exit_credit=750.0,
        vix=18.5, delta=0.42,
        group_id="2026-04-23-093000-580C", dte=0,
    )
    base.update(overrides)
    return base


class TestAppendRows:
    def test_writes_hardcoded_and_formulas(self):
        wb, ws = _seed_journal_with_row2()
        trades = [_sample_trade(), _sample_trade(strike=585.0, group_id="2026-04-23-093100-585C")]
        js.append_rows(ws, trades, start_row=3, last_trade_num=10)

        # Row 3: Trade # = 11, correct hardcoded values.
        assert ws.cell(3, js.COL["trade_num"]).value == 11
        assert ws.cell(3, js.COL["date"]).value == date(2026, 4, 23)
        assert ws.cell(3, js.COL["strike"]).value == 580.0
        assert ws.cell(3, js.COL["vwap"]).value == "Above"

        # Row-substituted formulas land correctly.
        assert ws.cell(3, js.COL["pl_dollar"]).value == "=U3+T3"
        assert ws.cell(3, js.COL["cum_pl"]).value == "=SUM($V$2:V3)"
        assert ws.cell(4, js.COL["pl_dollar"]).value == "=U4+T4"

    def test_does_not_touch_existing_rows(self):
        wb, ws = _seed_journal_with_row2()
        ws.cell(2, 1, value="EXISTING")
        js.append_rows(ws, [_sample_trade()], start_row=3, last_trade_num=0)
        assert ws.cell(2, 1).value == "EXISTING"

    def test_aa_ae_remain_blank(self):
        wb, ws = _seed_journal_with_row2()
        js.append_rows(ws, [_sample_trade()], start_row=3, last_trade_num=0)
        for col_idx in range(27, 32):  # AA–AE
            assert ws.cell(3, col_idx).value is None

    def test_row_height_set(self):
        wb, ws = _seed_journal_with_row2()
        js.append_rows(ws, [_sample_trade()], start_row=3, last_trade_num=0)
        assert ws.row_dimensions[3].height == js.ROW_HEIGHT

    def test_alignment_left_on_notes_cols_center_elsewhere(self):
        wb, ws = _seed_journal_with_row2()
        js.append_rows(ws, [_sample_trade()], start_row=3, last_trade_num=0)
        # AA (27) should be left-aligned.
        assert ws.cell(3, 27).alignment.horizontal == "left"
        # Strike (6) should be center-aligned.
        assert ws.cell(3, 6).alignment.horizontal == "center"


# ───────────────────────── extend_table_range ─────────────────────────

class TestExtendTableRange:
    def test_updates_table_end_row(self):
        from openpyxl.worksheet.table import Table

        wb = Workbook()
        ws = wb.active
        ws.title = js.SHEET_NAME
        for c in range(1, 38):
            ws.cell(row=1, column=c, value=f"H{c}")
        for c in range(1, 38):
            ws.cell(row=2, column=c, value=None)

        tbl = Table(displayName="Table2", ref="A1:AK2")
        ws.add_table(tbl)

        js.extend_table_range(ws, new_last_row=50)
        assert ws.tables["Table2"].ref == "A1:AK50"
